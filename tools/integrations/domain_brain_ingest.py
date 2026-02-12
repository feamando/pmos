import io
import os
import re
import sys
from pathlib import Path

import yaml
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

try:
    import openpyxl

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print(
        "Warning: openpyxl not installed. Install with: pip install openpyxl",
        file=sys.stderr,
    )

# Add tools directory to path for config_loader
sys.path.insert(0, str(Path(__file__).parent.parent))
import config_loader

# Config
SHEET_ID = "1DwsNzmbIYucZ65ehMZ74KDo_wMTX03yPfq4JNyZcrto"
ROOT_PATH = config_loader.get_root_path()
USER_PATH = ROOT_PATH / "user"
# Points to gdrive_mcp credentials
TOKEN_FILE = str(USER_PATH / ".secrets" / "gdrive_mcp" / "token.json")
SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]
BRAIN_DIR = USER_PATH / "brain"
ENTITIES_DIR = str(BRAIN_DIR / "Entities")
STRATEGY_DIR = str(BRAIN_DIR / "Strategy")
REGISTRY_FILE = str(BRAIN_DIR / "registry.yaml")

# Ensure directories exist
os.makedirs(ENTITIES_DIR, exist_ok=True)
os.makedirs(STRATEGY_DIR, exist_ok=True)


def get_drive_service():
    if not os.path.exists(TOKEN_FILE):
        print(f"Error: Token file not found at {TOKEN_FILE}")
        return None
    creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)
    return build("drive", "v3", credentials=creds)


def download_sheet_as_xlsx(service, file_id):
    print(f"Downloading sheet {file_id} as XLSX...")
    request = service.files().export_media(
        fileId=file_id,
        mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    fh = io.BytesIO()
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while not done:
        status, done = downloader.next_chunk()
        print(f"Download {int(status.progress() * 100)}%")
    fh.seek(0)
    return fh


def clean_name(name):
    if not name:
        return ""
    return str(name).strip()


def safe_filename(name):
    return (
        clean_name(name)
        .replace(" ", "_")
        .replace("&", "and")
        .replace("/", "_")
        .replace(":", "")
        .replace("\n", "_")
    )


def process_dos_tab(ws, registry_updates):
    print(f"Processing Tab: {ws.title}")

    # Identify headers row
    header_row_idx = None
    headers = {}

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3:
            print(f"Row {i}: {row}")
        # Look for specific columns to identify header
        row_str = [str(c).lower() if c else "" for c in row]

        # Header row usually has 'domain' and 'driver' or 'owner'
        if "domain" in row_str and ("driver" in row_str or "owner" in row_str):
            header_row_idx = i + 1
            for col_idx, cell in enumerate(row):
                if not cell:
                    continue
                c_val = str(cell).lower().strip()
                if c_val == "domain":
                    headers["domain"] = col_idx
                if "central functions" in c_val:
                    headers["alliance"] = col_idx
                if "driver" in c_val:
                    headers["comm_leader"] = col_idx
                if "yp owner" in c_val or "2026-yp owner" in c_val:
                    headers["tech_leader"] = col_idx
                # Multiple approvers, pick the last one or specific one
                if c_val == "approver":
                    # If we haven't found one, take it. If we found one (col 9), and find another (col 18), take 18 for YP?
                    # Let's take the one closer to YP columns for YP approver
                    headers["approver"] = col_idx
                if "refined" in c_val and "document" in c_val:
                    headers["do_doc"] = col_idx
                if "yp final" in c_val or ("final" in c_val and "market" in c_val):
                    headers["yp_doc"] = col_idx
            break

    if not header_row_idx:
        print("Could not find headers in DOs tab")
        return

    print(f"Found headers at row {header_row_idx}: {headers}")

    # Process rows
    current_alliance = None

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        domain_val = row[headers["domain"]] if "domain" in headers else None

        # Check if domain_val is an ID (number) and the next column is the name
        domain_name = domain_val
        if isinstance(domain_val, (int, float)) or (
            isinstance(domain_val, str) and domain_val.replace(".", "").isdigit()
        ):
            # Check next column
            next_col_val = row[headers["domain"] + 1]
            if next_col_val:
                domain_name = next_col_val

        alliance_val = row[headers["alliance"]] if "alliance" in headers else None

        # Handle merged cells or grouping (Alliance often spans rows)
        if alliance_val:
            current_alliance = clean_name(alliance_val)
            # Check if ID
            if isinstance(alliance_val, (int, float)) or (
                isinstance(alliance_val, str)
                and alliance_val.replace(".", "").isdigit()
            ):
                # Check next column
                next_col_val = row[headers["alliance"] + 1]
                if next_col_val:
                    current_alliance = clean_name(next_col_val)
        elif not alliance_val and current_alliance:
            # Assume continuation if domain is present
            pass

        if not domain_name:
            continue

        domain_name = clean_name(domain_name)
        if domain_name.lower() == "domain":
            continue  # header repeat?
        if domain_name == "None":
            continue

        # Extract other fields
        comm_leader = (
            clean_name(row[headers["comm_leader"]]) if "comm_leader" in headers else ""
        )
        tech_leader = (
            clean_name(row[headers["tech_leader"]]) if "tech_leader" in headers else ""
        )
        approver = clean_name(row[headers["approver"]]) if "approver" in headers else ""
        do_link = clean_name(row[headers["do_doc"]]) if "do_doc" in headers else ""
        yp_link = clean_name(row[headers["yp_doc"]]) if "yp_doc" in headers else ""

        # Create Domain Entity
        fname = f"Domain_{safe_filename(domain_name)}.md"
        fpath = os.path.join(ENTITIES_DIR, fname)

        content = f"""# Domain: {domain_name}

## Metadata
- **Type**: Domain
- **Alliance**: {current_alliance}
- **Commercial Leader**: [[{comm_leader.replace(chr(10), ', ')}]]
- **Tech Leader**: [[{tech_leader.replace(chr(10), ', ')}]]
- **Approver**: [[{approver.replace(chr(10), ', ')}]]

## Strategy Documents
- [Domain Objective Document]({do_link})
- [Yearly Plan Document]({yp_link})

## Context
Part of the {current_alliance} alliance.
"""
        with open(fpath, "w", encoding="utf-8") as f:
            f.write(content)

        registry_updates[domain_name] = {"type": "Domain", "file": f"Entities/{fname}"}
        print(f"  Created Domain: {domain_name}")

        # Register People
        for raw_p in [comm_leader, tech_leader, approver]:
            if not raw_p:
                continue
            # Split by newline or comma
            people = [p.strip() for p in re.split(r"[\n,]", raw_p) if p.strip()]

            for p in people:
                if len(p) > 2 and "TBD" not in p and "done" not in p.lower():
                    p_fname = f"{safe_filename(p)}.md"
                    p_path = os.path.join(ENTITIES_DIR, p_fname)
                    # Don't overwrite existing people (from squad ingestion) to preserve their context
                    # unless it's a stub? No, safer to not overwrite.
                    if not os.path.exists(p_path):
                        with open(p_path, "w", encoding="utf-8") as f:
                            f.write(
                                f"# {p}\n\n## Metadata\n- **Type**: Person\n- **Role**: Leadership\n"
                            )
                    registry_updates[p] = {
                        "type": "Person",
                        "file": f"Entities/{p_fname}",
                    }

        # Create/Update Yearly Plan Stub
        if current_alliance:
            yp_name = f"{current_alliance} Yearly Plan 2026"
            yp_fname = f"YP_2026_{safe_filename(current_alliance)}.md"
            yp_path = os.path.join(STRATEGY_DIR, yp_fname)

            # Simple check to avoid overwriting with less info, but appending links is good
            if not os.path.exists(yp_path):
                with open(yp_path, "w", encoding="utf-8") as f:
                    f.write(f"""# {yp_name}

## Metadata
- **Type**: Yearly Plan
- **Alliance**: {current_alliance}
- **Year**: 2026

## Links
- [Yearly Plan Document]({yp_link})
""")


def main():
    if not HAS_OPENPYXL:
        print(
            "Skipping domain_brain_ingest: openpyxl not installed. Install with: pip install openpyxl",
            file=sys.stderr,
        )
        sys.exit(0)  # Graceful exit - optional dependency

    service = get_drive_service()
    if not service:
        return

    xlsx_data = download_sheet_as_xlsx(service, SHEET_ID)
    wb = openpyxl.load_workbook(xlsx_data)

    registry_updates = {}

    # Find relevant tabs
    for sheet_name in wb.sheetnames:
        lower_name = sheet_name.lower()
        if "do" in lower_name and "configuration" in lower_name:
            process_dos_tab(wb[sheet_name], registry_updates)
        elif "tab 2" in lower_name:  # Fallback
            process_dos_tab(wb[sheet_name], registry_updates)

    # Update Registry
    print("Updating Registry...")
    if os.path.exists(REGISTRY_FILE):
        with open(REGISTRY_FILE, "r") as f:
            registry = yaml.safe_load(f) or {}
    else:
        registry = {}

    if "entities" not in registry:
        registry["entities"] = {}

    for name, data in registry_updates.items():
        if name not in registry["entities"]:
            registry["entities"][name] = data

    with open(REGISTRY_FILE, "w") as f:
        yaml.dump(registry, f, sort_keys=True)

    print("Done.")


if __name__ == "__main__":
    main()
